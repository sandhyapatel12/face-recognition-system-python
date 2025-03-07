import face_recognition # For face detection and recognition
import cv2 # OpenCV for accessing the webcam and image processing
import numpy as np # NumPy for handling arrays
import tkinter as tk
from tkinter import ttk, messagebox # For displaying pop-up messages
import pyttsx3 # For text-to-speech
import time # For handling time-based functions
import os
import openpyxl  # used to read, write, and edit Excel files (.xlsx) without needing Excel software
from datetime import datetime
import threading   # for open fastly the camera(without this camera take too much time for open)
from tqdm import tqdm
from PyQt6.QtWidgets import QApplication, QMessageBox
import sys
from openpyxl.drawing.image import Image  # Import Image module



# --------------------- INITIALIZED QApplication ----------------------

app = QApplication(sys.argv)

# --------------------- PYTTSX3 INITIALIZED ----------------------

# Initialize text-to-speech engine
engine = pyttsx3.init()

# --------------------- CAMERA INITIALIZED ----------------------

# Use DirectShow for Windows (Removes delay)
video_capture = cv2.VideoCapture(0, cv2.CAP_DSHOW) #DirectShow is a Windows-specific method that reduces camera delay.
#here 0 means the default camera (your laptop's built-in webcam). the webcam refers to the video capture device (your laptop or external camera) that is used to capture real-time video.)

# Check if the camera opened successfully
if not video_capture.isOpened():
    print("❌ Error: Could not open the camera.")
    exit()

# Set camera resolution for **faster processing**
video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)   # video frame height and width   (we can also set dynamic size)
video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)


# **Preload Frames for Faster Startup**
def warmup_camera():
    print("🔄 Warming up camera...")

    # underscore (_) is used as a throwaway variable in Python .It means, "I don’t care about this value."(here i don't want to this i am also creating for stabilize the camera so use it. it just ignore no more meaning )
    for _ in range(5):  # Captures 5 temporary frames which is allow the camera to stabilize before using it. (like your phone camera—it takes a second to adjust brightness and focus.)
        success, _ = video_capture.read()
        if not success:
            print("❌ Error: Camera frame not captured during warm-up.")
            exit()
    print("✅ Camera is ready!")

# **Run Warmup in a Separate Thread**
# Using threads makes programs faster and more responsive, especially when dealing with tasks like camera initialization.
warmup_thread = threading.Thread(target=warmup_camera)
warmup_thread.start() # runs in background
warmup_thread.join()  # Waits for completion before proceeding

print("🚀 Camera initialized instantly!")

# ---------------------- LOAD KNOWN FACES ----------------------

# Load images of known people and encode their faces
sandhya_img = face_recognition.load_image_file("faces/sandhya.png") # load image in memory

# face_encodings() detects the face in the image and converts it into a numerical representation that can be used for face recognition.
sandhya_encoding = face_recognition.face_encodings(sandhya_img)[0]  # [0] is used to select the first detected face in the image.( because an image may contain multiple faces.)
# print(sandhya_encoding)  #print this is We will see a long list of numbers (a NumPy array),  and select [0] index from this list

radha_img = face_recognition.load_image_file("faces/radha.jpg")
radha_encoding = face_recognition.face_encodings(radha_img)[0]  # Extract face features

# Store encodings and names of known faces in lists
known_face_encodings = [sandhya_encoding, radha_encoding]  # Encoded face data
known_face_names = ["Sandhya", "radha"]  # Names corresponding to the encodingsx


# ---------------------- ATTENDANCE FILE SETUP IN EXCEL ----------------------

current_date = datetime.now().strftime('%d-%m-%Y')
excel_file = f"{current_date}_attendance.xlsx"   # automatically create excel file which name also current date

# Function to create or load the Excel file
def get_attendance_workbook():

    # Check if the Excel file already exists
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)   # If the file exists, open and load the existing workbook
        #load_workbook is a function in the openpyxl library that opens (loads) an existing Excel file (.xlsx). It allows you to read, modify, and save the file.

    else:
        workbook = openpyxl.Workbook()         # If the file does not exist, create a new Excel workbook
        sheet = workbook.active #Gets the currently active sheet in the Excel file
        sheet.append(["Name",  "Time", "Screenshot"])  # Add header in excel file
        workbook.save(excel_file)  # Save the newly created Excel file
    return workbook     # Return the workbook object (either newly created or loaded)

# ---------------------- START VIDEO CAPTURE ----------------------

detection_period = 10   # set face detection time is 10 second
countdown_time = 5   # after on the camera that give us to 5 second for stabilize(prepare for video)

def main():
    try:
        start_time = time.time() #stores the current time in the start_time variable.
        capturing = True  # capturing process should start

        # when capturing process is running
        while True: #while true creates an infinite loop, meaning the program will keep running until manually stopped or a break statement is executed.
        
            _, frame = video_capture.read() # tells the camera to take a picture (frame) continuously.
            # The camera gives two things:
            # A success/fail message (True or False) → We don't need this, so we use _ to ignore it.
            # The actual picture (frame) → This is stored in frame.
            
            #  make the camera window full screen when it opens.
            cv2.namedWindow("Attendance", cv2.WND_PROP_FULLSCREEN)   #Attendance  is a window name
            cv2.setWindowProperty("Attendance", cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN) 
            #setWindowProperty  tells the window to open in full screen mode.
            
            # find how much time has gone by
            passed_time = time.time() - start_time #time.time() gives the current time in seconds  and start_time is the time when the countdown started

            # time still remaining and store into time_left variable
            # Ensure that the remaining time never goes below 0 (to avoid negative values)
            time_left = max(0, countdown_time - passed_time) #countdown_time - passed_time calculates how many seconds are left.



            # ---------------------- COUNTDOWN DISPLAY SCREEN----------------------

            #Check if the countdown is still running.
            if passed_time < countdown_time:
                    
                    # Define the circle's center and radius
                    center_x, center_y = frame.shape[1] // 2, 100  # Positioned at the top center
                    radius = 50  

                    # Calculate the angle for the countdown (360° means full time)
                    angle = int((time_left / countdown_time) * 360)

                    # Draw a circular outline (gray)
                    cv2.circle(frame, (center_x, center_y), radius, (100, 100, 100), 5)

                    # Draw the countdown progress (Green Arc)
                    cv2.ellipse(frame, (center_x, center_y), (radius, radius), -90, 0, angle, (0, 255, 0), 5)  # (0, 255, 0) is Green

                    # Display countdown text inside the circle
                    font = cv2.FONT_HERSHEY_DUPLEX
                    cv2.putText(frame, f"{int(time_left)}s", (center_x - 20, center_y + 10), font, 1, (255, 255, 255), 2)

                    # Show updated frame
                    cv2.imshow("Attendance", frame)
                    cv2.waitKey(1)
                    continue

            break  # Exit countdown loop

        start_time = time.time() #This records the current time when face recognition starts.

        # while capturing: means the program continuously captures frames from the camera until a face is properly detected or some stopping condition is met.
        while capturing:
            _, frame = video_capture.read() #_ is used because cv2.read() returns two values, but the first one (a boolean flag) is ignored. and frame contains the image from the video feed.

            # Resize the frame to a smaller size because face detection works faster on smaller images.  
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)  # here frame come from above , fx=0.25, fy=0.25 scales both width and height, (0, 0) means OpenCV will automatically calculate the new dimensions based on the scaling factors (fx, fy).  

            # OpenCV reads images in BGR(blue, green ,red) format, but the face_recognition library needs RGB(red, green,blue) format.
            rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB) #BGR AND RGB are color models used in image processing:

            #It detects faces in the given image (rgb_small_frame) and returns their positions (coordinates).
            face_locations = face_recognition.face_locations(rgb_small_frame)
            # print(face_locations)
            #If one face is detected, the output [(50, 150, 100, 100)]  # (top, right, bottom, left)
            #If multiple faces are detected: [(50, 150, 100, 100), (200, 300, 250, 250)]  
            #If no faces are found: [] empty list

            #Converts detected faces into numerical face encodings (a unique vector for each face).This is used to compare faces later.
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            detected_name = None #Later, this variable will store the name of the recognized person.

            workbook = get_attendance_workbook()   # call the function which is created on above
            sheet = workbook.active # Open the first page (active sheet)

            # compares detected faces with known faces and records attendance 
            for face_encoding in face_encodings: # check each detected face one by one.

                #Compare the detected face with all known faces.
                # known_face_encodings store all faces which stored by us in a list
                # face_encoding is the encoding of the detected face from the current frame.
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)

                # Calculate how different the detected face is from each stored face.
                # Smaller values mean a better match.
                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)

                # Get the index of the best match (face with the smallest distance).
                best_match_index = np.argmin(face_distances)

                # If a match is found, get the person’s name from the known faces list.
                match_found = matches[best_match_index]  # Check if there is a match
                matched_name = known_face_names[best_match_index]  # Store the matched name

                if match_found:
                    detected_name = matched_name

                else:
                  detected_name = "Unknown" 
                    
                
                    # Check If the Person is Already Marked Present
                already_present = any(row[0] == detected_name for row in sheet.iter_rows(min_row=2, values_only=True)) #It checks the Excel sheet to see if the name is already there.

                if already_present:
                    detected_name = "Already Present"

                else:

               
                    # Ensure a valid filename (replace ':' with '-')
                    current_time = datetime.now().strftime("%H:%M:%S")  # Use '-' instead of ':'
                    screenshot_file = f"{detected_name}_{current_date}.png"

            
                    # Capture the frame and ensure it's successfully saved
                    success = cv2.imwrite(screenshot_file, frame)

                    if success:
                        sheet.append([detected_name, current_time, screenshot_file])
                        workbook.save(excel_file)
                    else:
                        print("❌ Failed to save screenshot. Check if 'frame' is captured correctly.")

                    
                capturing = False  # Stop capturing
                break  # Exit loop after detecting a face

                
                

            # ---------------------- DISPLAY TIME LEFT SCREEN----------------------

            # find passed time means how much time is passed
            passed_time = time.time() - start_time # time.time() display current time and start time display that time when 10s timer is start  this through we can get passed time

            # find remaining time
            time_left = max(0, detection_period - passed_time)    # above we set detection_period for 10s   and 0 means  time left is never negative means not  go in - value

            # extracts the width and height of the video frame to adjust the progress bar accordingly.
            frame_height, frame_width, _ = frame.shape  #OpenCV, frame.shape returns three values(height, width, channels) here i don't need thrird value so use _ means ignore it

            # progessBar properties (Full width at the top of the screen)
            x, y = 0, 0  # Start at the top-left corner
            width, height = frame_width, 50  # Full width, height = 40 pixels
            filled_width = int((1 - (time_left / detection_period)) * width)  # Filling progress

            # Use a single black color for the progress bar
            bar_color = (0, 0, 0)  # Black color for a sleek look

            # Create a semi-transparent overlay
            overlay = frame.copy() # Make a copy of the original frame
            cv2.rectangle(overlay, (x, y), (x + width, y + height), (50, 50, 50, 100), -1)  # Dark transparent background
            cv2.rectangle(overlay, (x, y), (x + filled_width, y + height), bar_color, -1)  # Draw the progress bar

            # Blend overlay with the original frame for transparency effect
            alpha = 0.6  # overlay is 60% visible and blends smoothly with the frame.
            cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0, frame)

            # Display time left in the middle of the progress bar
            cv2.putText(frame, f"{int(time_left)}s", (width // 2 - 20, height - 10),
                        cv2.FONT_HERSHEY_DUPLEX, 1, (255, 255, 255), 2)

            cv2.imshow("Attendance", frame) #displays the updated frame with the progress bar on the screen.


            # ---------------------- EXIT OR STOP THE PROCESS----------------------

            # if the user press q key then stop capturing and exit loop
            if cv2.waitKey(1) & 0xFF == ord("q"):  
                #cv2.waitkey(1) This function waits for 1 millisecond and checks if a key is pressed.
                #If no key is pressed, it returns -1. If a key is pressed, it returns the key’s ASCII value.
                #ord("q") returns the ASCII value of "q", which is 113.
                #& 0xFF ensures that we get only the last 8 bits of the key’s ASCII value.

                capturing = False
                break

            # if time  is finished means over 10s
            if passed_time > detection_period:
                detected_name = "Timeout"
                capturing = False
                break

        # ---------------------- DIALOG BOX FOR DISPLAY MESSAGE----------------------


        def show_custom_dialog(title, message, icon="info", retry = False):
            
            msg_box = QMessageBox() #create msg box
            msg_box.setWindowTitle(title) #set  title
            msg_box.setText(message) # set msg

            # Set the appropriate icon
            if icon == "info":
                msg_box.setIcon(QMessageBox.Icon.Information)
            elif icon == "warning":
                msg_box.setIcon(QMessageBox.Icon.Warning)
            elif icon == "error":
                msg_box.setIcon(QMessageBox.Icon.Critical)
            else:
                msg_box.setIcon(QMessageBox.Icon.NoIcon)

            # used when face is not detected (for try again)
            if retry:
                msg_box.setStandardButtons(QMessageBox.StandardButton.Retry | QMessageBox.StandardButton.Cancel)

            # other wise display only ok button    
            else:
                msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)

            result = msg_box.exec()  # pauses the program until the user clicks a button (e.g., "OK", "Retry", "Cancel").and Once the user makes a choice, exec() returns the button they clicked

            # If the user clicks "Retry," restart the process
            if retry and result == QMessageBox.StandardButton.Retry:
                main()  # Restart face recognition process (Ensure 'main' function is properly defined)

        # ----------- dialog boxes for all cases -----------

        if detected_name in known_face_names:
            message = f"Attendance marked successfully for {detected_name}!"
            engine.say(message)  #convert text into speech
            engine.runAndWait()  #  waits for the speech to finish before continuing with the next lines of code.
            show_custom_dialog("Attendance", message, "info")

        elif detected_name == "Timeout":
            message = "Timeout Face not detected. Try again?"
            engine.say(message)
            engine.runAndWait()  
            show_custom_dialog("Attendance", message, "warning", retry=True)

        elif detected_name == "Already Present":
            message = "You are already present"
            engine.say(message)
            engine.runAndWait() 
            show_custom_dialog("Attendance", message, "info")

        elif detected_name == "Unknown":  
            message = " You are not part of the company."
            engine.say(message)
            engine.runAndWait()  
            show_custom_dialog("Attendance", message, "error")

        else:
            message = "Unexpected error occurred."
            engine.say(message)
            engine.runAndWait()
            show_custom_dialog("Attendance", message, "error")    

        # engine.runAndWait()



    # ---------------------- Automatically close the camera after processing ----------------------

        video_capture.release() #Stops the camera and frees up all the resources  which in uses like(camera, ram, cpu)
        cv2.destroyAllWindows() # Closes all OpenCV windows that may have opened.
        # print("✅ Camera closed successfully after attendance processing!")

    # If any error happens, the program won’t crash , it catches the error and prints the error message.
    except Exception as e:
        print(f"❌ An error occurred: {e}")

    #The finally block always runs, whether there's an error or not.
    finally:  #double-checks everything is properly closed if not close then close forcefully
        if video_capture.isOpened():
            video_capture.release()
            cv2.destroyAllWindows()
        print("✅ Camera forcefully closed in finally block!")

# This ensures that the script only runs when executed directly(means run when 'main.py' only this not), not when imported as a module in another script(file code).
if __name__ == "__main__":
    main()        
